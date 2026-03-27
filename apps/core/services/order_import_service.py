import csv
import io
from decimal import Decimal, InvalidOperation

from dateutil import parser as date_parser
from django.db import transaction

from ..models import FinanceTransaction, Order, OrderItem, Reservation, SKU


class OrderImportService:
    TEMPLATE_HEADERS = [
        '客户昵称',
        '手机号码',
        '地址',
        '款式',
        '客订来源',
        '租金',
        '租金渠道',
        '预收押金',
        '押金渠道',
        '预定时间',
        '发货时间',
        '距离发货剩余天数',
        '状态',
        '发货单号',
        '经手人',
        '操作时间',
        '备注',
    ]

    HEADER_ALIASES = {
        'customer_name': ['客户昵称', '客户姓名', '昵称', '姓名'],
        'customer_phone': ['手机号码', '手机号', '联系电话'],
        'delivery_address': ['地址', '收货地址', '配送地址'],
        'style': ['款式', 'SKU', 'SKU名称', '产品名称', '产品编号', 'SKU编号'],
        'order_source': ['客订来源', '来源', '来源平台', '订单来源'],
        'rental_amount': ['租金', '租金金额'],
        'rental_channel': ['租金渠道', '尾款渠道'],
        'deposit_paid': ['预收押金', '押金', '押金金额'],
        'deposit_channel': ['押金渠道'],
        'event_date': ['预定时间', '预定日期', '活动日期', '预定单时间'],
        'ship_date': ['发货时间', '发货日期'],
        'ship_tracking': ['发货单号', '快递单号', '物流单号'],
        'status': ['状态', '订单状态'],
        'operator': ['经手人', '负责人', '客服'],
        'operation_time': ['操作时间', '登记时间', '录入时间'],
        'notes': ['备注', '说明'],
        'source_order_no': ['平台单号', '订单平台单号'],
        'customer_wechat': ['微信号', '客户微信', '微信'],
        'xianyu_order_no': ['闲鱼订单号'],
        'quantity': ['数量', '套数'],
        'return_tracking': ['回寄单号', '回收单号'],
        'record_type': ['订单类型', '单据类型', '记录类型', '类型'],
    }

    SOURCE_MAP = {
        '微信': 'wechat',
        '微信成交': 'wechat',
        '闲鱼': 'xianyu',
        '小红书': 'xiaohongshu',
        '小程序': 'miniprogram',
        '其他': 'other',
        'wechat': 'wechat',
        'xianyu': 'xianyu',
        'xiaohongshu': 'xiaohongshu',
        'miniprogram': 'miniprogram',
        'other': 'other',
    }

    ORDER_STATUS_MAP = {
        '待处理': 'pending',
        '待发货': 'confirmed',
        '已发货': 'delivered',
        '使用中': 'in_use',
        '已归还': 'returned',
        '已完成': 'completed',
        '已取消': 'cancelled',
        'pending': 'pending',
        'confirmed': 'confirmed',
        'delivered': 'delivered',
        'in_use': 'in_use',
        'returned': 'returned',
        'completed': 'completed',
        'cancelled': 'cancelled',
    }

    RESERVATION_STATUS_MAP = {
        '待补信息': 'pending_info',
        '待确认': 'pending_info',
        '待跟进': 'pending_info',
        '待联系': 'pending_info',
        '可转正式订单': 'ready_to_convert',
        '可转单': 'ready_to_convert',
        '待转单': 'ready_to_convert',
        '可下正式单': 'ready_to_convert',
        '已转订单': 'converted',
        '已转单': 'converted',
        'converted': 'converted',
        '已取消': 'cancelled',
        '取消预定': 'cancelled',
        'reservation_cancelled': 'cancelled',
        '已退款': 'refunded',
        '已退订金': 'refunded',
        'refunded': 'refunded',
        'pending_info': 'pending_info',
        'ready_to_convert': 'ready_to_convert',
        'cancelled': 'cancelled',
    }

    NOTE_FIELDS = ['rental_channel', 'deposit_channel', 'operator', 'operation_time']

    @staticmethod
    def _normalize_text(value):
        return str(value or '').strip()

    @staticmethod
    def _decode_bytes(content):
        for encoding in ('utf-8-sig', 'utf-8', 'gb18030'):
            try:
                return content.decode(encoding)
            except UnicodeDecodeError:
                continue
        return content.decode('latin1')

    @staticmethod
    def _load_rows(uploaded_file):
        filename = (getattr(uploaded_file, 'name', '') or '').lower()
        content = uploaded_file.read()
        if filename.endswith('.xlsx'):
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise ValueError('当前环境未安装 openpyxl，暂时无法读取 .xlsx，请先安装依赖或改传 CSV UTF-8。') from exc
            workbook = load_workbook(io.BytesIO(content), data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                raise ValueError('导入文件为空')
            headers = [str(item or '').strip() for item in rows[0]]
            return [
                {headers[index]: ('' if value is None else value) for index, value in enumerate(row)}
                for row in rows[1:]
            ], headers

        text = OrderImportService._decode_bytes(content)
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise ValueError('导入文件缺少表头')
        return list(reader), list(reader.fieldnames)

    @staticmethod
    def _normalize_header_text(value):
        return str(value or '').replace('\ufeff', '').replace('\u3000', ' ').strip().lower()

    @staticmethod
    def _normalize_header_map(headers):
        mapped = {}
        for raw in headers:
            header = str(raw or '').replace('\ufeff', '').replace('\u3000', ' ').strip()
            if not header:
                continue
            normalized_header = OrderImportService._normalize_header_text(header)
            for canonical, aliases in OrderImportService.HEADER_ALIASES.items():
                if any(normalized_header == OrderImportService._normalize_header_text(alias) for alias in aliases):
                    mapped[canonical] = header
                    break
        return mapped

    @staticmethod
    def _get_value(row, header_map, key):
        header = header_map.get(key)
        if not header:
            return ''
        return str(row.get(header, '') or '').strip()

    @staticmethod
    def _parse_decimal(value, default='0'):
        raw = str(value or '').strip().replace('￥', '').replace(',', '')
        if not raw:
            return Decimal(default)
        try:
            return Decimal(raw)
        except InvalidOperation as exc:
            raise ValueError(f'金额格式不正确：{value}') from exc

    @staticmethod
    def _parse_int(value, default=1):
        raw = str(value or '').strip()
        if not raw:
            return default
        try:
            return int(float(raw))
        except Exception as exc:
            raise ValueError(f'数量格式不正确：{value}') from exc

    @staticmethod
    def _parse_date(value, field_label):
        raw = str(value or '').strip()
        if not raw:
            return None
        try:
            return date_parser.parse(raw).date()
        except Exception as exc:
            raise ValueError(f'{field_label}格式不正确：{value}') from exc

    @staticmethod
    def _resolve_sku(style_value):
        style = OrderImportService._normalize_text(style_value)
        if not style:
            raise ValueError('缺少产品名称')

        exact = SKU.objects.filter(is_active=True).filter(name__iexact=style).first()
        if exact:
            return exact
        raise ValueError(f'产品管理中不存在名称为“{style}”的产品')

    @staticmethod
    def _resolve_record_type(row, header_map):
        record_type = OrderImportService._normalize_text(OrderImportService._get_value(row, header_map, 'record_type')).lower()
        if record_type in {'预定单', '预订单', '预定', '预约', '意向单', 'reservation'}:
            return 'reservation'
        if record_type in {'订单', '正式订单', 'order'}:
            return 'order'

        status_raw = OrderImportService._normalize_text(OrderImportService._get_value(row, header_map, 'status'))
        if status_raw in OrderImportService.RESERVATION_STATUS_MAP:
            return 'reservation'
        return 'order'

    @staticmethod
    def _build_notes(row, header_map):
        notes = []
        base_notes = OrderImportService._get_value(row, header_map, 'notes')
        if base_notes:
            notes.append(base_notes)

        for key in OrderImportService.NOTE_FIELDS:
            value = OrderImportService._get_value(row, header_map, key)
            if value:
                label = OrderImportService.HEADER_ALIASES[key][0]
                notes.append(f'{label}：{value}')

        consumed_headers = {header for header in header_map.values()}
        extra_pairs = []
        for raw_key, raw_value in row.items():
            key = str(raw_key or '').strip()
            value = str(raw_value or '').strip()
            if not key or not value or key in consumed_headers:
                continue
            extra_pairs.append(f'{key}={value}')
        if extra_pairs:
            notes.append('导入扩展字段：' + '；'.join(extra_pairs))

        return '\n'.join(notes)

    @staticmethod
    def _normalize_row(row, header_map):
        record_type = OrderImportService._resolve_record_type(row, header_map)
        sku = OrderImportService._resolve_sku(OrderImportService._get_value(row, header_map, 'style'))
        customer_name = OrderImportService._get_value(row, header_map, 'customer_name')
        customer_phone = OrderImportService._get_value(row, header_map, 'customer_phone')
        delivery_address = OrderImportService._get_value(row, header_map, 'delivery_address')
        if not customer_name or not customer_phone or not delivery_address:
            raise ValueError('客户昵称、手机号码、地址不能为空')

        event_date = OrderImportService._parse_date(
            OrderImportService._get_value(row, header_map, 'event_date'),
            '预定时间',
        )
        if not event_date:
            raise ValueError('预定时间不能为空')

        ship_date = OrderImportService._parse_date(
            OrderImportService._get_value(row, header_map, 'ship_date'),
            '发货时间',
        )
        quantity = max(OrderImportService._parse_int(
            OrderImportService._get_value(row, header_map, 'quantity'),
            default=1,
        ), 1)
        rental_amount = OrderImportService._parse_decimal(
            OrderImportService._get_value(row, header_map, 'rental_amount'),
            default=str(sku.rental_price),
        )
        deposit_paid = OrderImportService._parse_decimal(
            OrderImportService._get_value(row, header_map, 'deposit_paid'),
            default='0',
        )
        source_raw = OrderImportService._get_value(row, header_map, 'order_source')
        order_source = OrderImportService.SOURCE_MAP.get(source_raw, 'wechat')
        status_raw = OrderImportService._get_value(row, header_map, 'status')
        notes = OrderImportService._build_notes(row, header_map)
        total_amount = rental_amount * quantity
        customer_wechat = OrderImportService._get_value(row, header_map, 'customer_wechat')
        if not customer_wechat and record_type == 'reservation':
            customer_wechat = customer_phone

        if record_type == 'reservation':
            status = OrderImportService.RESERVATION_STATUS_MAP.get(status_raw, 'pending_info')
            balance = Decimal('0.00')
        else:
            status = OrderImportService.ORDER_STATUS_MAP.get(status_raw, 'pending')
            balance = Decimal('0.00') if status in ['delivered', 'in_use', 'returned', 'completed'] else total_amount

        return {
            'record_type': record_type,
            'customer_name': customer_name,
            'customer_phone': customer_phone,
            'customer_wechat': customer_wechat,
            'xianyu_order_no': OrderImportService._get_value(row, header_map, 'xianyu_order_no'),
            'order_source': order_source,
            'source_order_no': OrderImportService._get_value(row, header_map, 'source_order_no'),
            'delivery_address': delivery_address,
            'return_address': delivery_address,
            'event_date': event_date,
            'rental_days': 1,
            'ship_date': ship_date,
            'ship_tracking': OrderImportService._get_value(row, header_map, 'ship_tracking'),
            'return_tracking': OrderImportService._get_value(row, header_map, 'return_tracking'),
            'total_amount': total_amount,
            'deposit_paid': deposit_paid,
            'balance': balance,
            'status': status,
            'notes': notes,
            'sku': sku,
            'quantity': quantity,
            'rental_price': rental_amount,
            'deposit_per_item': deposit_paid if quantity == 1 else (deposit_paid / quantity if deposit_paid else Decimal('0.00')),
            'deposit_amount': deposit_paid,
            'city': '',
        }

    @staticmethod
    def preview_file(uploaded_file, default_sku_id=None, preview_limit=20):
        rows, headers = OrderImportService._load_rows(uploaded_file)
        header_map = OrderImportService._normalize_header_map(headers)
        required_keys = ['customer_name', 'customer_phone', 'delivery_address', 'style', 'event_date']
        missing = [OrderImportService.HEADER_ALIASES[key][0] for key in required_keys if key not in header_map]
        if missing:
            raise ValueError(f'导入文件缺少必要列：{"、".join(missing)}')
        preview_rows = []
        errors = []
        valid_count = 0

        for index, row in enumerate(rows, start=2):
            if not any(str(value or '').strip() for value in row.values()):
                continue
            try:
                normalized = OrderImportService._normalize_row(row, header_map)
                valid_count += 1
                if len(preview_rows) < preview_limit:
                    preview_rows.append({
                        'row_no': index,
                        'record_type_label': '预定单' if normalized['record_type'] == 'reservation' else '正式订单',
                        'customer_name': normalized['customer_name'],
                        'customer_phone': normalized['customer_phone'],
                        'sku_name': normalized['sku'].name,
                        'event_date': str(normalized['event_date']),
                        'status_label': (
                            dict(Reservation.STATUS_CHOICES).get(normalized['status'], normalized['status'])
                            if normalized['record_type'] == 'reservation'
                            else dict(Order.STATUS_CHOICES).get(normalized['status'], normalized['status'])
                        ),
                        'total_amount': normalized['total_amount'],
                    })
            except Exception as exc:
                row_hint = ' / '.join([
                    str(row.get(header_map.get('customer_name', ''), '') or '').strip(),
                    str(row.get(header_map.get('customer_phone', ''), '') or '').strip(),
                ]).strip(' /')
                errors.append(f'第 {index} 行（{row_hint or "未识别客户"}）：{exc}')

        return {
            'headers': headers,
            'preview_rows': preview_rows,
            'valid_count': valid_count,
            'error_count': len(errors),
            'errors': errors,
            'total_rows': valid_count + len(errors),
        }

    @staticmethod
    @transaction.atomic
    def import_file(uploaded_file, user, default_sku_id=None):
        rows, headers = OrderImportService._load_rows(uploaded_file)
        header_map = OrderImportService._normalize_header_map(headers)
        required_keys = ['customer_name', 'customer_phone', 'delivery_address', 'style', 'event_date']
        missing = [OrderImportService.HEADER_ALIASES[key][0] for key in required_keys if key not in header_map]
        if missing:
            raise ValueError(f'导入文件缺少必要列：{"、".join(missing)}')
        created_orders = []
        created_reservations = []
        errors = []

        for index, row in enumerate(rows, start=2):
            if not any(str(value or '').strip() for value in row.values()):
                continue
            try:
                normalized = OrderImportService._normalize_row(row, header_map)
                if normalized['record_type'] == 'reservation':
                    reservation = Reservation.objects.create(
                        customer_wechat=normalized['customer_wechat'],
                        customer_name=normalized['customer_name'],
                        customer_phone=normalized['customer_phone'],
                        city=normalized['city'],
                        delivery_address=normalized['delivery_address'],
                        sku=normalized['sku'],
                        quantity=normalized['quantity'],
                        event_date=normalized['event_date'],
                        deposit_amount=normalized['deposit_amount'],
                        status=normalized['status'],
                        notes=normalized['notes'],
                        created_by=user,
                        owner=user,
                    )
                    if normalized['deposit_amount'] > Decimal('0.00'):
                        FinanceTransaction.objects.create(
                            reservation=reservation,
                            transaction_type='reservation_deposit_received',
                            amount=normalized['deposit_amount'],
                            notes='导入预定单收取订金',
                            created_by=user,
                        )
                    created_reservations.append(reservation)
                else:
                    order = Order.objects.create(
                        customer_name=normalized['customer_name'],
                        customer_phone=normalized['customer_phone'],
                        customer_wechat=normalized['customer_wechat'],
                        xianyu_order_no=normalized['xianyu_order_no'],
                        order_source=normalized['order_source'],
                        source_order_no=normalized['source_order_no'],
                        delivery_address=normalized['delivery_address'],
                        return_address=normalized['return_address'],
                        event_date=normalized['event_date'],
                        rental_days=normalized['rental_days'],
                        ship_date=normalized['ship_date'],
                        ship_tracking=normalized['ship_tracking'],
                        return_tracking=normalized['return_tracking'],
                        total_amount=normalized['total_amount'],
                        deposit_paid=normalized['deposit_paid'],
                        balance=normalized['balance'],
                        status=normalized['status'],
                        notes=normalized['notes'],
                        created_by=user,
                    )
                    OrderItem.objects.create(
                        order=order,
                        sku=normalized['sku'],
                        quantity=normalized['quantity'],
                        rental_price=normalized['rental_price'],
                        deposit=normalized['deposit_per_item'],
                        subtotal=normalized['total_amount'],
                    )
                    created_orders.append(order)
            except Exception as exc:
                row_hint = ' / '.join([
                    str(row.get(header_map.get('customer_name', ''), '') or '').strip(),
                    str(row.get(header_map.get('customer_phone', ''), '') or '').strip(),
                ]).strip(' /')
                errors.append(f'第 {index} 行（{row_hint or "未识别客户"}）：{exc}')

        return {
            'created_orders': created_orders,
            'created_reservations': created_reservations,
            'errors': errors,
            'created_count': len(created_orders) + len(created_reservations),
            'created_order_count': len(created_orders),
            'created_reservation_count': len(created_reservations),
            'error_count': len(errors),
        }
